from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename='sample.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
print(ws['A1'].value)

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
ws['A2'] = '2'
print(ws['A2'].value)

print(int(ws['A1'].value)+int(ws['A2'].value))
# Save the file
#wb.save("sample.xlsx")

for row in ws.rows:
    for cell in row:
        print(cell.value)
        
